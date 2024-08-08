from flask import Flask, request, redirect, url_for, render_template, Response
import pandas as pd
from openpyxl import load_workbook
import csv
import os
import pyexcel as pe
from xlsx2csv import Xlsx2csv

app = Flask(__name__, template_folder='templates')


@app.route(rule='/', methods=['GET', 'POST'])
def index():
    if request.method == 'GET':
        return render_template('index.html')
    elif request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        if username == 'admin' and password == 'admin':
            return 'Success'
        else:
            return 'Fail'


@app.route('/file_upload', methods=['POST'])
def file_upload():
    file = request.files['file']

    if file.content_type == 'text/csv':
        df = pd.read_csv(file)
    elif file.content_type in ['application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']:
        df = pd.read_excel(file)
    elif file.content_type == 'text/plain':
        df = pd.read_csv(file)
    elif file.content_type == 'application/json':
        df = pd.read_json(file)
    elif file.content_type == 'application/octet-stream':
        df = pd.read_csv(file)
    else:
        return 'File format not supported'

    return df.to_html()


@app.route('/convert_csv', methods=['POST'])
def convert_csv():
    file = request.files['file']
    df = pd.read_excel(file)
    response = Response(df.to_csv(), mimetype='text/csv', headers={
        'Content-Disposition': 'attachment;filename=converted.csv'
    })
    return response


@app.route('/convert_csv_two', methods=['POST'])
def convert_csv_two():
    file = request.files['file']
    df = pd.read_excel(file)
    response = Response(df.to_csv(), mimetype='text/csv', headers={
        'Content-Disposition': 'attachment;filename=converted.csv'
    })
    return response


# Yöntem 1: pandas ile Dönüştürme
@app.route('/upload_file_pandas', methods=['POST'])
def upload_file_pandas():
    if 'file' not in request.files:
        return redirect(request.url)
    file = request.files['file']
    if file.filename == '':
        return redirect(request.url)

    df = pd.read_excel(file)
    csv_path = os.path.join('uploads', 'converted_file_pandas.csv')
    df.to_csv(csv_path, index=False)

    return f'File converted and saved to {csv_path} using pandas.'


# Yöntem 2: openpyxl ve csv modülleri ile Dönüştürme
@app.route('/upload_file_openpyxl', methods=['POST'])
def upload_file_openpyxl():
    if 'file' not in request.files:
        return redirect(request.url)
    file = request.files['file']
    if file.filename == '':
        return redirect(request.url)

    wb = load_workbook(file)
    sheet = wb.active
    csv_path = os.path.join('uploads', 'converted_file_openpyxl.csv')

    with open(csv_path, 'w', newline='') as f:
        c = csv.writer(f)
        for r in sheet.iter_rows(values_only=True):
            c.writerow(r)

    return f'File converted and saved to {csv_path} using openpyxl.'


# Yöntem 3: pyexcel ile Dönüştürme
@app.route('/upload_file_pyexcel', methods=['POST'])
def upload_file_pyexcel():
    if 'file' not in request.files:
        return redirect(request.url)
    file = request.files['file']
    if file.filename == '':
        return redirect(request.url)

    sheet = pe.get_sheet(file_type="xlsx", file_content=file.read())
    csv_path = os.path.join('uploads', 'converted_file_pyexcel.csv')
    sheet.save_as(csv_path)

    return f'File converted and saved to {csv_path} using pyexcel.'


# Yöntem 4: xlsx2csv ile Dönüştürme
@app.route('/upload_file_xlsx2csv', methods=['POST'])
def upload_file_xlsx2csv():
    if 'file' not in request.files:
        return redirect(request.url)
    file = request.files['file']
    if file.filename == '':
        return redirect(request.url)

    csv_path = os.path.join('uploads', 'converted_file_xlsx2csv.csv')
    Xlsx2csv(file, outputencoding="utf-8").convert(csv_path)

    return f'File converted and saved to {csv_path} using xlsx2csv.'


if __name__ == '__main__':
    if not os.path.exists('uploads'):
        os.makedirs('uploads')
    app.run(host='0.0.0.0', port=5555, debug=True)
